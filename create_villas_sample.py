import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Villas"

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY        = "1A3E6F"   # standard header bg
BRIGHT_BLUE = "1E4D8C"   # required-column header bg
WHITE       = "FFFFFF"
ALT_BLUE    = "EEF4FF"
NOTE_GREY   = "888888"

# ── Fills ─────────────────────────────────────────────────────────────────────
fill_navy   = PatternFill("solid", fgColor=NAVY)
fill_bright = PatternFill("solid", fgColor=BRIGHT_BLUE)
fill_white  = PatternFill("solid", fgColor=WHITE)
fill_alt    = PatternFill("solid", fgColor=ALT_BLUE)

# ── Fonts ─────────────────────────────────────────────────────────────────────
font_header = Font(name="Arial", size=11, bold=True,  color=WHITE)
font_data   = Font(name="Arial", size=11, bold=False, color="000000")
font_note   = Font(name="Arial", size=10, italic=True, color=NOTE_GREY)

# ── Alignment ─────────────────────────────────────────────────────────────────
align_center = Alignment(horizontal="center", vertical="center")
align_left   = Alignment(horizontal="left",   vertical="center")

# ── Border (thin all around) ──────────────────────────────────────────────────
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Headers (row 1) ───────────────────────────────────────────────────────────
headers = [
    ("A", "VillaNumber",    True),
    ("B", "OwnerName",      True),
    ("C", "PhoneNumber",    True),
    ("D", "Area",           False),
    ("E", "Annual Fee",     False),
    ("F", "Deposit Amount", False),
    ("G", "Password",       False),
    ("H", "2024 Debt",      False),
    ("I", "2025 Debt",      False),
]

for col_letter, label, required in headers:
    cell = ws[f"{col_letter}1"]
    cell.value     = label
    cell.font      = font_header
    cell.fill      = fill_bright if required else fill_navy
    cell.alignment = align_center
    cell.border    = border

ws.row_dimensions[1].height = 32

# ── Data rows (2-6) ───────────────────────────────────────────────────────────
data = [
    ("A-01", "أحمد محمد علي",      "+201001234567", 250, 18000, 10000, "123456",  0,    0),
    ("A-02", "محمد إبراهيم حسن",    "+201112345678", 300, 24000, 15000, "123456",  5000, 0),
    ("B-01", "سارة أحمد الخطيب",    "+201223456789", 275, 18000,     0, "123456",  0,    3000),
    ("B-02", "عمر خالد النجار",     "+201334567890", 320, 24000, 20000, "123456",  2000, 1500),
    ("C-01", "فاطمة علي مصطفى",    "+201445678901", 200, 18000, 10000, "123456",  0,    0),
]

for i, row_data in enumerate(data, start=2):
    fill = fill_white if i % 2 == 0 else fill_alt
    for j, value in enumerate(row_data, start=1):
        cell = ws.cell(row=i, column=j)
        cell.value     = value
        cell.font      = font_data
        cell.fill      = fill
        cell.alignment = align_center
        cell.border    = border
    ws.row_dimensions[i].height = 24

# ── Row 7: empty, still needs borders + fill ─────────────────────────────────
for col in range(1, 10):
    cell = ws.cell(row=7, column=col)
    cell.border = border
    cell.fill   = fill_white

# ── Row 8: merged note ────────────────────────────────────────────────────────
ws.merge_cells("A8:I8")
note_cell = ws["A8"]
note_cell.value = (
    "* Required: VillaNumber, OwnerName, PhoneNumber  |  "
    "Default password: 123456  |  "
    "Annual Fee: 18000 or 24000  |  "
    "Phone must include country code e.g. +20"
)
note_cell.font      = font_note
note_cell.alignment = Alignment(horizontal="left", vertical="center")
note_cell.border    = border
ws.row_dimensions[8].height = 24

# Apply border to the merged cells too (openpyxl only needs the top-left cell
# for merged regions, but we ensure the border wraps the whole merge region)
for col in range(2, 10):
    ws.cell(row=8, column=col).border = border

# ── Number / text formats ─────────────────────────────────────────────────────
TEXT_FMT   = "@"
NUMBER_FMT = '#,##0'

# PhoneNumber (C) → text
for row in range(1, 9):
    ws.cell(row=row, column=3).number_format = TEXT_FMT

# Password (G) → text
for row in range(1, 9):
    ws.cell(row=row, column=7).number_format = TEXT_FMT

# Number columns: D=4, E=5, F=6, H=8, I=9
for col in (4, 5, 6, 8, 9):
    for row in range(2, 7):
        ws.cell(row=row, column=col).number_format = NUMBER_FMT

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {"A": 15, "B": 28, "C": 22, "D": 12,
              "E": 15, "F": 18, "G": 15, "H": 15, "I": 15}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ── Freeze pane ───────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = r"E:\NEW APP\villas_sample.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
